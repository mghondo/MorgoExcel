import os
import openpyxl
from datetime import datetime
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.worksheet.page import PageMargins

lib_folder = os.getcwd()
file_render_folder = os.path.join(lib_folder, 'MORNINGDROP')
file_complete_folder = os.path.join(lib_folder, 'MORNINGCOMPLETE')

os.makedirs(file_render_folder, exist_ok=True)
os.makedirs(file_complete_folder, exist_ok=True)

def process_morning_file(input_path, location=None):
    current_date = datetime.now().strftime("%m.%d.%Y")
    
    try:
        print(f"Input file: {input_path}")
        
        workbook = openpyxl.load_workbook(input_path)
        sheet = workbook.active

        sheet.delete_rows(1, 4)
        sheet.delete_cols(9, 2)
        col_h_values = [cell.value for cell in sheet['H']]
        sheet.delete_cols(8)
        sheet.insert_cols(1)
        for i, value in enumerate(col_h_values, start=1):
            sheet.cell(row=i, column=1, value=value)
        sheet.delete_cols(6)

        rows_to_keep = []
        for row in sheet.iter_rows(min_row=1, values_only=False):
            if row[4].value != "Gear":
                rows_to_keep.append(row)

        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active

        header_row = rows_to_keep[0]
        new_sheet.print_title_rows = '1:1'
        data_rows = rows_to_keep[1:]

        for row in data_rows:
            cell = row[3]
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value[-4:]

        sorted_data_rows = sorted(data_rows, key=lambda x: (x[0].value.lower(), int(x[3].value)))

        new_sheet.append([cell.value for cell in header_row])

        for row in sorted_data_rows:
            new_sheet.append([cell.value for cell in row])

        new_sheet.cell(row=1, column=4, value="Last 4 PKG ID")
        new_sheet.cell(row=1, column=2, value="Product Name")
        new_sheet.cell(row=1, column=8, value="Fulfillment")
        new_sheet.cell(row=1, column=9, value="Vault")
        new_sheet.cell(row=1, column=10, value="Quarantine")
        new_sheet.cell(row=1, column=11, value="Total")
        new_sheet.cell(row=1, column=12, value="\u2713")

        new_sheet.oddHeader.left.text = "Discrepancies: _____________________________"
        # Include location in the header
        header_text = f"{location} Morning Count {current_date}" if location else f"Morning Count {current_date}"
        new_sheet.oddHeader.center.text = header_text
        new_sheet.oddHeader.right.text = "Name + Badge: _____________________________"
        new_sheet.oddFooter.center.text = "&P"

        new_sheet.page_margins = PageMargins(left=0, right=0)
        new_sheet.page_setup.orientation = 'landscape'

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in new_sheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.font = Font(bold=True, size=10)

        new_sheet.freeze_panes = 'A2'

        for cell in new_sheet[1]:
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        new_sheet.print_gridlines = True

        # Use the provided location for the filename
        location_prefix = location if location else "Temp"
        output_filename = f"{location_prefix}{current_date}.MorningCount.xlsx"
        output_path = os.path.join(file_complete_folder, output_filename)
        
        new_sheet.delete_cols(5)

        new_workbook.save(output_path)
        print(f"Processed file saved as: {output_filename}")
        print(f"File saved to: {output_path}")
        
        return output_filename

    except Exception as e:
        print(f"An error occurred while processing the Excel file: {str(e)}")
        print(f"Input file path: {input_path}")
        print(f"Current working directory: {os.getcwd()}")
        raise