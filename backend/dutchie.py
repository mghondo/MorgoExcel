import os
import csv
import openpyxl
import datetime

def process_dutchie_file(input_file, output_file):
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Read the CSV file
        with open(input_file, 'r', newline='') as csvfile:
            csv_reader = csv.reader(csvfile)
            headers = next(csv_reader)  # Read the header row

            # Find the index of the "Vendor", "Category", "Room", and "Package ID" columns
            vendor_index = headers.index("Vendor")
            category_index = headers.index("Category")
            room_index = headers.index("Room")

            # Find Package ID index case-insensitively
            package_id_index = next((i for i, h in enumerate(headers) if h.strip().lower() == "package id"), None)

            # Move column "Vendor" to the far left (column A)
            new_headers = [headers[vendor_index]] + headers[:vendor_index] + headers[vendor_index+1:]
            ws.append(new_headers)  # Append the modified header row to the worksheet

            backstock_rows = []
            quarantine_rows = []
            other_rows = []

            for row in csv_reader:
                # Skip rows where Category is "Gear"
                if row[category_index].strip().lower() == "gear":
                    continue

                # Move column "Vendor" to the far left (column A)
                new_row = [row[vendor_index]] + row[:vendor_index] + row[vendor_index+1:]

                # Modify Package ID to get the last four digits
                new_package_id_index = next((i for i, h in enumerate(new_headers) if h.strip().lower() == "package id"), None)
                if new_package_id_index is not None:
                    new_row[new_package_id_index] = new_row[new_package_id_index][-4:]

                # Separate rows based on Room value
                if row[room_index].strip().lower() == "backstock":
                    backstock_rows.append(new_row)
                elif row[room_index].strip().lower() == "quarantine":
                    quarantine_rows.append(new_row)
                else:
                    other_rows.append(new_row)

            # Combine all rows
            all_rows = backstock_rows + quarantine_rows + other_rows

            # Sort rows based on Room, Vendor, and Package ID
            def sort_key(row, headers):
                # Adjust indices since "Vendor" is moved to the first position
                vendor_index = 0
                room_index = next((i for i, h in enumerate(headers) if h.strip().lower() == "room"), None)
                package_id_index = next((i for i, h in enumerate(headers) if h.strip().lower() == "package id"), None)
                
                # Try to convert Package ID to int, otherwise use a high value to place at the end
                try:
                    package_id_value = int(row[package_id_index])
                except ValueError:
                    package_id_value = float('inf')  # Place non-numeric IDs at the end
                
                # Sort key
                return (row[room_index].lower(), row[vendor_index].lower(), package_id_value)

            # Apply sorting
            sorted_rows = sorted(all_rows, key=lambda row: sort_key(row, new_headers))

            # Append sorted rows to the worksheet
            for row in sorted_rows:
                ws.append(row)

            # Delete columns "Category", "Tags", and "Strain"
            ws.delete_cols(new_headers.index("Strain") + 1)
            ws.delete_cols(new_headers.index("Tags") + 1)
            ws.delete_cols(new_headers.index("Category") + 1)

            # Delete all columns starting with G and going to the right
            ws.delete_cols(7, ws.max_column - 6)

            # Add columns titled "Fulfillment", "Vault", "Quarantine", "Backstock", "Total" and then Checkmark Symbol in order in the columns to the right
            ws.insert_cols(7, 4)
            ws.cell(row=1, column=7, value="Fulfillment")
            ws.cell(row=1, column=8, value="Vault")
            ws.cell(row=1, column=9, value="Quarantine")
            ws.cell(row=1, column=10, value="Backstock")
            ws.cell(row=1, column=11, value="Total")
            ws.cell(row=1, column=12, value="✔")

        current_date = datetime.date.today().strftime("%m-%d-%Y")

        ws.oddHeader.left.text = "Discrepancies: _____________________________"
        ws.oddHeader.center.text = f"Dutchie Weekly {current_date}"
        ws.oddHeader.right.text = "Name + Badge: _____________________________"
        ws.oddFooter.center.text = "&P"

        # Determine the output filename based on the value in cell B2
        if ws.cell(row=2, column=2).value and ws.cell(row=2, column=2).value.startswith('M'):
            # output_file = os.path.join("DUTCHIE-OUT", f'Marengo-Dutchie-{datetime.date.today().strftime("%Y-%m-%d")}.xlsx')
            output_file = os.path.join("DUTCHIE-OUT", f'Marengo-Dutchie-{datetime.date.today().strftime("%m-%d-%Y")}.xlsx')

        else:
            # output_file = os.path.join("DUTCHIE-OUT", f'Columbus-Dutchie-{datetime.date.today().strftime("%Y-%m-%d")}.xlsx')
            output_file = os.path.join("DUTCHIE-OUT", f'Columbus-Dutchie-{datetime.date.today().strftime("%m-%d-%Y")}.xlsx')

        # Save the workbook as XLSX
        wb.save(output_file)
        print(f"Successfully saved: {output_file}")
    except Exception as e:
        print(f"Error processing file {input_file}: {str(e)}")

def clear_output_directory(output_dir):
    for filename in os.listdir(output_dir):
        if filename != 'temp':
            file_path = os.path.join(output_dir, filename)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
                    print(f"Deleted: {file_path}")
            except Exception as e:
                print(f"Failed to delete {file_path}. Reason: {e}")

def main():
    input_dir = "DUTCHIE-IN"
    output_dir = "DUTCHIE-OUT"

    print(f"Input directory: {os.path.abspath(input_dir)}")
    print(f"Output directory: {os.path.abspath(output_dir)}")

    # Create output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Clear the output directory
    clear_output_directory(output_dir)
    print("Cleared DUTCHIE-OUT directory (except 'temp' file)")

    # Get current date in YYYY-MM-DD format
    current_date = datetime.date.today().strftime("%Y-%m-%d")

    # Process all CSV files in the input directory
    files_processed = 0
    for filename in os.listdir(input_dir):
        if filename.endswith('.csv'):
            input_path = os.path.join(input_dir, filename)
            output_filename = f'DUTCHIE-{current_date}.xlsx'
            output_path = os.path.join(output_dir, output_filename)
            process_dutchie_file(input_path, output_path)
            print(f"Processed: {filename} -> {output_filename}")
            files_processed += 1

    if files_processed == 0:
        print("No CSV files found in the input directory.")
    else:
        print(f"Total files processed: {files_processed}")

if __name__ == "__main__":
    main()
