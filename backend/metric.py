import os
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

def process_metric_file(input_file, output_file):
    try:
        # Verify input file exists
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"Input file not found: {input_file}")

        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Read the CSV file
        print(f"Reading CSV file: {input_file}")
        with open(input_file, 'r', newline='', encoding='utf-8') as csvfile:
            csv_reader = csv.reader(csvfile)
            rows = list(csv_reader)  # Load all rows to verify structure
            if not rows:
                raise ValueError("CSV file is empty")
            for row in rows:
                ws.append(row)

        # Verify worksheet has data
        if ws.max_row < 1 or ws.max_column < 1:
            raise ValueError("Worksheet is empty after loading CSV")

        # Move data from column S to T
        print("Moving data from column S to T")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=19, max_col=19):
            for cell in row:
                ws.cell(row=cell.row, column=20, value=cell.value)

        # Delete columns S, P to K, and H to C
        print("Deleting columns S, P to K, and H to C")
        ws.delete_cols(19)  # Delete column S
        ws.delete_cols(11, 6)  # Delete columns K to P
        ws.delete_cols(3, 6)  # Delete columns C to H

        # Rename column B
        if ws['B1'].value == "Item":
            print("Renaming column B from 'Item' to 'Description'")
            ws['B1'].value = "Description"
            ws['B1'].alignment = openpyxl.styles.Alignment(wrap_text=True)

        # Rename column C
        if ws['C1'].value == "Quantity":
            print("Renaming column C from 'Quantity' to 'METRIC\\nQuantity'")
            ws['C1'].value = "METRIC\nQuantity"
            ws['C1'].alignment = openpyxl.styles.Alignment(wrap_text=True)

        # Delete the last three columns (E, F, G)
        print("Deleting last three columns")
        if ws.max_column >= 3:
            ws.delete_cols(ws.max_column - 2, 3)
        else:
            print("Warning: Not enough columns to delete last three")

        # Add page number in the middle footer
        ws.oddFooter.center.text = "Page &[Page]"

        # Add current date in the middle header
        current_date = datetime.now().strftime("%Y-%m-%d")
        ws.oddHeader.center.text = f"Metric {current_date}"

        # Ensure output directory exists
        output_dir = os.path.dirname(output_file)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # Save the workbook as XLSX
        print(f"Saving output file: {output_file}")
        wb.save(output_file)

        # Verify output file was created
        if not os.path.exists(output_file):
            raise FileNotFoundError(f"Output file was not created: {output_file}")

    except Exception as e:
        print(f"Error in process_metric_file: {str(e)}")
        raise

def clear_output_directory(output_dir):
    try:
        # Delete all files in the output directory except 'temp'
        for filename in os.listdir(output_dir):
            if filename != 'temp':
                file_path = os.path.join(output_dir, filename)
                try:
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                        print(f"Deleted file: {file_path}")
                except Exception as e:
                    print(f"Failed to delete {file_path}. Reason: {e}")
    except Exception as e:
        print(f"Error in clear_output_directory: {str(e)}")
        raise

def main():
    input_dir = "METRIC-IN"
    output_dir = "METRIC-OUT"

    # Create output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    else:
        # Clear the output directory
        clear_output_directory(output_dir)

    print("Cleared METRIC-OUT directory (except 'temp' file)")

    # Get current date in YYYY-MM-DD format
    current_date = datetime.date.today().strftime("%m-%d-%Y")

    # Process all CSV files in the input directory
    for filename in os.listdir(input_dir):
        if filename.endswith('.csv'):
            input_path = os.path.join(input_dir, filename)
            output_filename = f'METRIC-{current_date}.xlsx'
            output_path = os.path.join(output_dir, output_filename)
            try:
                process_metric_file(input_path, output_path)
                print(f"Processed: {filename} -> {output_filename}")
            except Exception as e:
                print(f"Failed to process {filename}: {str(e)}")

if __name__ == "__main__":
    main()