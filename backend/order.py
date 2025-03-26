import os
from openpyxl import load_workbook
import math

def process_order_file(file_path, num_of_days):
    """Process order file and return vendor data"""
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        sheet.delete_rows(1, 4)  # Remove first four rows
        
        vendor_data = {}
        column_indices = {
            'ProductDesc': None,
            'VendorName': None,
            'PosQty': None,
            'RemainingQty': None
        }

        # Identify column indices from headers
        for cell in sheet[1]:
            if cell.value in column_indices:
                column_indices[cell.value] = cell.column - 1

        # Process data rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                product_desc = row[column_indices['ProductDesc']]
                vendor_name = row[column_indices['VendorName']]
                pos_qty = float(row[column_indices['PosQty']])
                remaining_qty = float(row[column_indices['RemainingQty']])
                
                if pos_qty > 0 and remaining_qty >= 0:
                    daily_sales = pos_qty / num_of_days
                    days_until_sold_out = math.ceil(remaining_qty / daily_sales) if daily_sales > 0 else float('inf')

                    if days_until_sold_out <= 14:
                        vendor_data.setdefault(vendor_name, []).append({
                            'name': product_desc,
                            'daysUntilSoldOut': days_until_sold_out
                        })
            
            except (TypeError, ValueError) as e:
                continue  # Skip invalid rows

        return vendor_data

    except Exception as e:
        raise Exception(f"Order processing error: {str(e)}")
